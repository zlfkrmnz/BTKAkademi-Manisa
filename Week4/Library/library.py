import sqlite3 as sql
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, date, timedelta


class Library:
    def __init__(self):
        self.db = sql.connect('C:/Users/zlfkr/PycharmProjects/BTKAkademi-Manisa/DataBase/database.sqlite3')
        self.cursor = self.db.cursor()

    def __del__(self):
        self.db.commit()
        self.db.close()

    def add_book(self, isbn: int, title: str, author: str, p_date: str, genre: str, summary: str):
        while True:
            x = input("Emin misiniz? (Evet/Hayır)")
            if x.lower() == "evet":
                command = "INSERT INTO books (ISBN, title, author, p_date, genre, summary) VALUES (?, ?, ?, ?, ?, ?)"
                self.cursor.execute(command, (isbn, title, author, p_date, genre, summary))
                print("Kitap başarıyla eklendi.")
                self.db.commit()
                break
            elif x.lower() == "hayır":
                print("İşlem iptal edildi.")
                break
            else:
                print("Lütfen geçerli bir seçenek giriniz.")

    def search_book(self, option, text):
        text = '%' + text + '%'
        command_isbn = "SELECT * FROM books WHERE ISBN LIKE ?"
        command_title = "SELECT * FROM books WHERE title LIKE ?"
        command_author = "SELECT * FROM books WHERE author LIKE ?"
        if option == "1":
            values = self.cursor.execute(command_isbn, (text,)).fetchall()
        elif option == "2":
            values = self.cursor.execute(command_title, (text,)).fetchall()
        elif option == "3":
            values = self.cursor.execute(command_author, (text,)).fetchall()
        else:
            return None
        if len(values) == 0:
            return "Herhangi bir kitap bulunamadı."
        else:
            self.db.commit()
            return values

    def delete(self, table_name):
        command = "SELECT * FROM " + table_name + " WHERE ISBN=?"
        while True:
            isbn = input("ISBN: ")
            values = self.cursor.execute(command, (isbn,)).fetchall()
            if len(values) == 0:
                print("Herhangi bir kitap bulunamadı. Doğru ISBN girdiğinizden emin olunuz.")
                x = input("Tekrar denemek için herhangi bir değer giriniz. Çıkmak için 'q' tuşuna basınız: ")
                if x.lower() == "q":
                    break
                else:
                    continue
            else:
                x = input("Emin misiniz? (Evet/Hayır)")
                if x.lower() == "evet":
                    try:
                        command = "DELETE FROM " + table_name + " WHERE ISBN=?"
                        self.cursor.execute(command, (isbn,))
                    except:
                        print("İşlem sırasında bir hata oluştu. Lütfen tekrar deneyiniz.")
                    self.db.commit()
                    print("İşlem başarıyla gerçekleştirildi.")
                    break
                elif x.lower() == "hayır":
                    print("İşlem iptal edildi.")
                    break
                else:
                    print("Lütfen geçerli bir seçenek giriniz.")
        return isbn

    def list_book(self):
        command = "SELECT * FROM books"
        values = self.cursor.execute(command).fetchall()
        if len(values) != 0:
            print(" KİTAPLAR ".center(174, "*"))
            for no, value in enumerate(values):
                print(f"Kitap No: {no + 1}")
                print(f"ISBN: {value[0]}\nKitap Adı: {value[1]}\nYazar: {value[2]}\nYayın Tarihi: {value[3]}"
                      f"\nTür: {value[4]}\nÖzet: {value[5]}\nKitap Kütüphanede mi?: {'Evet' if (value[6] == 'true') else 'Hayır'}")
                print("*" * 174)
        else:
            print("Herhangi bir kitap bulunamadı.")

    def borrow_book(self):
        command = "SELECT inLibrary FROM books WHERE ISBN=?"
        while True:
            text = input("ISBN: ")
            value = self.cursor.execute(command, (text,)).fetchone()[0]
            if value is None:
                print("Bu ISBN ile kayıtlı herhangi bir kitap kütüphanede yer almamaktadır.")
                x = input("Tekrar denemek için herhangi bir değer giriniz. Çıkmak için 'q' tuşuna basınız: ")
                if x.lower() == "q":
                    break
                else:
                    continue
            elif value == "true":
                x = input("Emin misiniz? (Evet/Hayır): ")
                if x.lower() == "evet":
                    try:
                        command = "INSERT INTO borrows (ISBN, borrow_date, who, return_date) VALUES (?, ?, ?, ?)"
                        who = input("Ad Soyad: ")
                        self.cursor.execute(command, (text, date.today().strftime("%d.%m.%Y"), who,
                                                      (date.today() + timedelta(days=15)).strftime("%d.%m.%Y")))
                        command = "UPDATE books SET inLibrary=? WHERE ISBN=?"
                        self.cursor.execute(command, ("false", text))
                        self.db.commit()
                        print("Ödünç alma işlemi başarıyla gerçekleşti.")
                        break
                    except:
                        print("Ödünç alma işlemi sırasında bir hata oluştu. Lütfen tekrar deneyiniz.")
                if x.lower() == "hayır":
                    print("İşlem iptal edildi.")
                    break
                else:
                    print("Lütfen geçerli bir seçenek giriniz.")
            elif value == "false":
                print("Ödünç almak istediğiniz kitap başkası tarafından ödünç alınmış. Lütfen daha sonra tekrar deneyin "
                      "veya başka bir kitap seçiniz.")
                x = input(
                    "Başka bir kitap ödünç almak için herhangi bir değer giriniz. Çıkmak için 'q' tuşuna basınız: ")
                if x.lower() == "q":
                    break
                else:
                    continue

    def return_book(self):
        try:
            text = self.delete("borrows")
            command = "UPDATE books SET inLibrary=? WHERE ISBN=?"
            self.cursor.execute(command, ("true", text))
            self.db.commit()
        except:
            print("Ödünç verme işlemi gerçekleştirilirken bir hata oluştu.")

    def backup(self):
        try:
            directories = os.listdir(os.chdir("Backup"))
            command = "SELECT * FROM books"
            books = self.cursor.execute(command).fetchall()
            command = "SELECT * FROM borrows"
            borrows = self.cursor.execute(command).fetchall()
            if len(directories) == 0:
                dir_name = datetime.now().strftime("%d.%m.%Y")
                os.mkdir(dir_name)
                os.chdir(dir_name)
                filename = datetime.now().strftime("%H.%M.%S") + ".xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = "books"
                for book in books:
                    ws.append(book)
                ws = wb.create_sheet("borrows")
                for borrow in borrows:
                    ws.append(borrow)
                wb.save(filename)
            else:
                dir_name = date.today().strftime("%d.%m.%Y")
                if dir_name in directories:
                    os.chdir(dir_name)
                    filename = datetime.now().strftime("%H.%M.%S") + ".xlsx"
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "books"
                    for book in books:
                        ws.append(book)
                    ws = wb.create_sheet("borrows")
                    for borrow in borrows:
                        ws.append(borrow)
                    wb.save(filename)
                else:
                    os.mkdir(dir_name)
                    os.chdir(dir_name)
                    filename = datetime.now().strftime("%H.%M.%S") + ".xlsx"
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "books"
                    for book in books:
                        ws.append(book)
                    ws = wb.create_sheet("borrows")
                    for borrow in borrows:
                        ws.append(borrow)
                    wb.save(filename)
            print("Veritabanı başarıyla yedeklendi.")
        except Exception as e:
            print("Yedekleme işlemi sırasında bir hata oluştu. ", e)
        os.chdir("..")
        os.chdir("..")

    def recovery(self):
        try:
            directories = os.listdir(os.chdir("Backup"))
            if len(directories) == 0:
                print("Herhangi bir yedek bulunamadı.")
            else:
                print("Yedek Listesi".center(50, "-"))
                for directory in directories:
                    print(directory)
                dir_name = input("Yedeklemeyi geri yüklemek istediğiniz klasörün adını giriniz: ")
                if dir_name in directories:
                    files = os.listdir(os.chdir(dir_name))
                    for file in files:
                        print(file)
                    file_name = input("Yedeklemeyi geri yüklemek istediğiniz dosyanın adını giriniz: ")
                    if file_name in files:
                        command = "DELETE FROM books"
                        self.cursor.execute(command)
                        command = "DELETE FROM borrows"
                        self.cursor.execute(command)
                        command = "INSERT INTO books (ISBN, title, author, p_date, genre, summary, inLibrary) VALUES (?, ?, ?, ?, ?, ?, ?)"
                        wb = load_workbook(file_name)
                        ws = wb.get_sheet_by_name("books")
                        for row in ws.iter_rows(values_only=True):
                            self.cursor.execute(command, row)
                        command = "INSERT INTO borrows (ISBN, borrow_date, who, return_date) VALUES (?, ?, ?, ?)"
                        ws = wb.get_sheet_by_name("borrows")
                        for row in ws.iter_rows(values_only=True):
                            self.cursor.execute(command, row)
                        print("Yedekleme veritabanına başarıyla geri yüklendi.")
                    else:
                        print("Girdiğiniz değer ile herhangi bir dosya eşleşmedi.")
                else:
                    print("Girdiğiniz değer ile herhangi bir klasör eşleşmedi.")
        except Exception as e:
            print("Yedeği veritabanına yüklerken hata oluştu. ", e)
        self.db.commit()
        os.chdir("..")
        os.chdir("..")
